#!/usr/bin/env python3
import sys
import json
import openpyxl
from copy import copy
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

def set_cell_value(ws, cell_ref, value):
    from openpyxl.utils import coordinate_to_tuple
    
    row, col = coordinate_to_tuple(cell_ref)
    
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            top_left = merged_range.start_cell
            ws[top_left.coordinate].value = value
            return
    
    try:
        ws[cell_ref].value = value
    except AttributeError:
        pass
    
def export_excel(data):
    template_path = '/home/runner/workspace/server/excel_export_template_1.xlsx'
    
    wb_data_only = openpyxl.load_workbook(template_path, data_only=True)
    ws_data = wb_data_only.active
    
    wb = openpyxl.load_workbook(template_path, keep_links=False)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                value_cell = ws_data[cell.coordinate]
                cell.value = value_cell.value
    
    calc = data['calculation']
    items = data['items']
    
    from datetime import datetime
    
    set_cell_value(ws, 'B2', calc.get('invoice_no', ''))
    
    invoice_date = calc.get('invoice_date', '')
    if invoice_date:
        try:
            if isinstance(invoice_date, str):
                date_obj = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
                formatted_date = date_obj.strftime('%d/%m/%Y')
            else:
                formatted_date = invoice_date
        except:
            formatted_date = invoice_date
    else:
        formatted_date = ''
    set_cell_value(ws, 'B4', formatted_date)
    
    set_cell_value(ws, 'C2', 'TOTAL VALUE')
    set_cell_value(ws, 'D2', float(calc.get('total_value', 0)))
    set_cell_value(ws, 'D4', int(calc.get('total_quantity', 0)))
    set_cell_value(ws, 'E2', 'TRANSPORT COST')
    set_cell_value(ws, 'E4', 'INSURANCE COST')
    set_cell_value(ws, 'G2', 'STORAGE COST')
    set_cell_value(ws, 'G4', 'CURRENCY -USD/TL-')
    set_cell_value(ws, 'F2', float(calc.get('transport_cost', 0)))
    set_cell_value(ws, 'F4', float(calc.get('insurance_cost', 0)))
    set_cell_value(ws, 'H2', float(calc.get('storage_cost', 0)))
    set_cell_value(ws, 'H4', float(calc.get('currency_rate', 0)))
    
    total_customs_tax = sum(float(item.get('customs_tax', 0)) for item in items)
    total_add_customs_tax = sum(float(item.get('additional_customs_tax', 0)) for item in items)
    total_kkdf = sum(float(item.get('kkdf', 0)) for item in items)
    total_vat = sum(float(item.get('vat', 0)) for item in items)
    total_tax_usd = sum(float(item.get('total_tax_usd', 0)) for item in items)
    total_tax_tl = sum(float(item.get('total_tax_tl', 0)) for item in items)
    
    set_cell_value(ws, 'A6', 'TOTAL CUSTOMS TAX')
    set_cell_value(ws, 'B6', 'TOTAL ADD. CUSTOMS TAX')
    set_cell_value(ws, 'C6', 'TOTAL KKDF')
    set_cell_value(ws, 'D6', 'TOTAL VAT (KKDF INCLUDED)')
    set_cell_value(ws, 'E6', 'TOTAL VAT (KKDF EXCLUDED)')
    set_cell_value(ws, 'F6', 'TOTAL TAX (KKDF INCLUDED)')
    set_cell_value(ws, 'G6', 'TOTAL TAX (KKDF EXCLUDED)')
    set_cell_value(ws, 'H6', 'TOTAL TAX TURKISH LIRA')
    
    set_cell_value(ws, 'A7', total_customs_tax)
    set_cell_value(ws, 'B7', total_add_customs_tax)
    set_cell_value(ws, 'C7', total_kkdf)
    set_cell_value(ws, 'D7', total_vat + total_kkdf)
    set_cell_value(ws, 'E7', total_vat)
    set_cell_value(ws, 'F7', total_tax_usd)
    set_cell_value(ws, 'G7', total_tax_usd - total_kkdf)
    set_cell_value(ws, 'H7', total_tax_tl)
    
    set_cell_value(ws, 'A9', 'HTS Codes')
    set_cell_value(ws, 'B9', 'Country of Origin')
    set_cell_value(ws, 'C9', 'Style')
    set_cell_value(ws, 'D9', 'Color')
    set_cell_value(ws, 'E9', 'Category')
    set_cell_value(ws, 'F9', 'Description')
    set_cell_value(ws, 'G9', 'Fabric Content')
    set_cell_value(ws, 'H9', 'Cost')
    set_cell_value(ws, 'I9', 'Unit')
    set_cell_value(ws, 'J9', 'Total Value')
    set_cell_value(ws, 'K9', 'TR HS CODE')
    set_cell_value(ws, 'L9', 'EX REGISTRY FORM')
    set_cell_value(ws, 'M9', 'AZO DYE TEST')
    set_cell_value(ws, 'N9', 'SPECIAL CUSTOMS')
    set_cell_value(ws, 'O9', 'TRANSPORT')
    set_cell_value(ws, 'P9', 'INSURANCE')
    set_cell_value(ws, 'Q9', 'STORAGE')
    set_cell_value(ws, 'R9', 'CUSTOMS TAX %')
    set_cell_value(ws, 'S9', 'ADDITIONAL CUSTOMS TAX %')
    set_cell_value(ws, 'T9', 'KKDF %')
    set_cell_value(ws, 'U9', 'VAT %')
    set_cell_value(ws, 'V9', 'TOTAL CUSTOMS TAX')
    set_cell_value(ws, 'W9', 'TOTAL ADDT CUSTOMS TAX')
    set_cell_value(ws, 'X9', 'KKDF VALUE')
    set_cell_value(ws, 'Y9', 'VAT BASE WITH KKDF')
    set_cell_value(ws, 'Z9', 'VAT BASE WITHOUT KKDF')
    set_cell_value(ws, 'AA9', 'VAT VALUE WITH KKDF')
    set_cell_value(ws, 'AB9', 'VAT VALUE WITHOUT KKDF')
    set_cell_value(ws, 'AC9', 'TOTAL TAX WITH KKDF')
    set_cell_value(ws, 'AD9', 'TOTAL TAX WITHOUT KKDF')
    
    template_row = 10
    last_data_row = 10 + len(items) - 1
    
    for index, item in enumerate(items):
        row_num = 10 + index
        
        if index > 0:
            ws.insert_rows(row_num)
            
            for col in range(1, 31):
                template_cell = ws.cell(template_row, col)
                new_cell = ws.cell(row_num, col)
                
                if template_cell.has_style:
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.fill = copy(template_cell.fill)
                    new_cell.number_format = copy(template_cell.number_format)
                    new_cell.protection = copy(template_cell.protection)
                    new_cell.alignment = copy(template_cell.alignment)
            
            ws.row_dimensions[row_num].height = ws.row_dimensions[template_row].height
            
            from openpyxl.worksheet.cell_range import CellRange
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == template_row and merged_range.max_row == template_row:
                    new_range = CellRange(
                        min_col=merged_range.min_col,
                        min_row=row_num,
                        max_col=merged_range.max_col,
                        max_row=row_num
                    )
                    ws.merge_cells(str(new_range))
        
        requirements = item.get('requirements', '')
        has_ex_registry = 'EX REGISTRY FORM' in requirements
        has_azo_dye = 'AZO DYE TEST' in requirements
        has_special = 'SPECIAL CUSTOM' in requirements
        
        from openpyxl.utils import get_column_letter
        
        set_cell_value(ws, f'A{row_num}', item.get('hts_code', ''))
        set_cell_value(ws, f'B{row_num}', item.get('country_of_origin', ''))
        set_cell_value(ws, f'C{row_num}', item.get('style', ''))
        set_cell_value(ws, f'D{row_num}', item.get('color', ''))
        set_cell_value(ws, f'E{row_num}', item.get('category', ''))
        set_cell_value(ws, f'F{row_num}', item.get('description', ''))
        set_cell_value(ws, f'G{row_num}', item.get('fabric_content', ''))
        set_cell_value(ws, f'H{row_num}', float(item.get('cost', 0)))
        set_cell_value(ws, f'I{row_num}', int(item.get('unit_count', 0)))
        set_cell_value(ws, f'J{row_num}', float(item.get('total_value', 0)))
        set_cell_value(ws, f'K{row_num}', item.get('tr_hs_code', ''))
        set_cell_value(ws, f'L{row_num}', 'X' if has_ex_registry else '')
        set_cell_value(ws, f'M{row_num}', 'X' if has_azo_dye else '')
        set_cell_value(ws, f'N{row_num}', 'X' if has_special else '')
        set_cell_value(ws, f'O{row_num}', float(item.get('transport_share', 0)))
        set_cell_value(ws, f'P{row_num}', float(item.get('insurance_share', 0)))
        set_cell_value(ws, f'Q{row_num}', float(item.get('storage_share', 0)))
        
        hs_code_data = item.get('hs_code_data')
        if hs_code_data:
            set_cell_value(ws, f'R{row_num}', float(hs_code_data.get('customs_tax_percent', 0)))
            set_cell_value(ws, f'S{row_num}', float(hs_code_data.get('additional_customs_tax_percent', 0)))
            set_cell_value(ws, f'T{row_num}', float(hs_code_data.get('kkdf_percent', 0)))
            set_cell_value(ws, f'U{row_num}', float(hs_code_data.get('vat_percent', 0)))
        else:
            set_cell_value(ws, f'R{row_num}', 0)
            set_cell_value(ws, f'S{row_num}', 0)
            set_cell_value(ws, f'T{row_num}', 0)
            set_cell_value(ws, f'U{row_num}', 0)
        
        item_customs_tax = float(item.get('customs_tax', 0))
        item_add_customs_tax = float(item.get('additional_customs_tax', 0))
        item_kkdf = float(item.get('kkdf', 0))
        item_vat = float(item.get('vat', 0))
        item_total_tax_usd = float(item.get('total_tax_usd', 0))
        
        vat_base_with_kkdf = float(item.get('vat_base', 0))
        vat_base_without_kkdf = vat_base_with_kkdf - item_kkdf
        
        vat_percent = float(hs_code_data.get('vat_percent', 0)) if hs_code_data else 0
        vat_without_kkdf = vat_base_without_kkdf * vat_percent
        
        set_cell_value(ws, f'V{row_num}', item_customs_tax)
        set_cell_value(ws, f'W{row_num}', item_add_customs_tax)
        set_cell_value(ws, f'X{row_num}', item_kkdf)
        set_cell_value(ws, f'Y{row_num}', vat_base_with_kkdf)
        set_cell_value(ws, f'Z{row_num}', vat_base_without_kkdf)
        set_cell_value(ws, f'AA{row_num}', item_vat)
        set_cell_value(ws, f'AB{row_num}', vat_without_kkdf)
        set_cell_value(ws, f'AC{row_num}', item_total_tax_usd)
        set_cell_value(ws, f'AD{row_num}', item_total_tax_usd - item_kkdf)
        
        gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        
        if index % 2 == 0:
            for col in range(1, 31):
                cell = ws.cell(row_num, col)
                cell.fill = gray_fill
        else:
            for col in range(1, 31):
                cell = ws.cell(row_num, col)
                cell.fill = white_fill
    
    white_side = Side(style='thin', color='FFFFFFFF')
    white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
    dxf = DifferentialStyle(border=white_border)
    
    # Range 1: A{last_data_row+1}:AD6000 (main data area below last row)
    start_row = last_data_row + 1
    fill_range_1 = f"A{start_row}:AD6000"
    
    for row in ws.iter_rows(min_row=start_row, max_row=min(ws.max_row, 6000), min_col=1, max_col=30):
        for cell in row:
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            if cell.value:
                cell.value = None
    
    rule_1 = Rule(type='expression', dxf=dxf, stopIfTrue=False, formula=['TRUE'])
    ws.conditional_formatting.add(fill_range_1, rule_1)
    
    # Range 2: AC1:BZ7 (top header area in extra columns)
    fill_range_2 = "AC1:BZ7"
    
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=29, max_col=78):
        for cell in row:
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            if cell.value:
                cell.value = None
    
    rule_2 = Rule(type='expression', dxf=dxf, stopIfTrue=False, formula=['TRUE'])
    ws.conditional_formatting.add(fill_range_2, rule_2)
    
    # Range 3: AE1:BZ6000 (large range in extra columns)
    fill_range_3 = "AE1:BZ6000"
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6000), min_col=31, max_col=78):
        for cell in row:
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            if cell.value:
                cell.value = None
    
    rule_3 = Rule(type='expression', dxf=dxf, stopIfTrue=False, formula=['TRUE'])
    ws.conditional_formatting.add(fill_range_3, rule_3)
    
    output_path = f"/tmp/tax_calculation_{calc.get('reference', 'export')}.xlsx"
    wb.save(output_path)
    
    return output_path

if __name__ == '__main__':
    data = json.loads(sys.stdin.read())
    output_path = export_excel(data)
    print(output_path)
