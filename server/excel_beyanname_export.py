#!/usr/bin/env python3
import sys
import json
import openpyxl
from openpyxl.styles import numbers
from copy import copy

# Country code mapping: 2-letter → 3-digit numeric code
DEFAULT_COUNTRY_CODE_MAPPING = {
    'CN': '720',
    'ID': '700',
    'KH': '696',
    'VN': '690',
    'US': '400',
    'TW': '736',
    'IT': '005',
    'RO': '066',
    'JO': '628',
    'NI': '432',
    'AQ': '891',
    'TH': '680',
    'LK': '669',
    'AL': '070',
    'SG': '706',
    'GT': '416',
    'CO': '480',
    'CM': '302',
    'PH': '708',
    'TR': '052',
    'CA': '404',
    'SV': '428',
    'HK': '740',
}

def export_beyanname_excel(data):
    """
    Export tax calculation data to BEYANNAME Excel template.
    Template structure:
    - Row 1: Headers with column descriptions
    - Row 2+: Product data
    """
    template_path = '/home/runner/workspace/server/EXCEL_BEYANNAME_AKTARIM.xlsx'
    
    # Load template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Merge default mappings with custom mappings
    custom_mappings = data.get('customMappings', {})
    country_code_mapping = {**DEFAULT_COUNTRY_CODE_MAPPING, **custom_mappings}
    
    items = data['items']
    
    # Start writing data from row 2 (row 1 is header)
    start_row = 2
    
    for index, item in enumerate(items):
        row_num = start_row + index
        
        # Get values with fallbacks
        tr_hs_code = item.get('tr_hs_code', '')
        cost = float(item.get('cost', 0))
        unit_count = int(item.get('unit_count', 0))
        total_value = cost * unit_count  # IMPORTANT: NOT CIF, just invoice value
        
        # HS Code data
        hs_code_data = item.get('hs_code_data', {})
        unit = hs_code_data.get('unit', '') if hs_code_data else ''
        description_tr = hs_code_data.get('description_tr', '') if hs_code_data else ''
        vat_percent_decimal = float(hs_code_data.get('vat_percent', 0)) if hs_code_data else 0
        vat_percent = vat_percent_decimal * 100  # Convert 0.1000 to 10
        
        # Product data (brand, item_description only)
        product_data = item.get('product_data', {})
        brand = product_data.get('brand', '') if product_data else ''
        item_description = product_data.get('item_description', '') if product_data else ''
        
        # Country of origin comes directly from the item (from tax_calculation_items table)
        # NOT from product_data - this ensures we use the country specified during this calculation
        country_of_origin = item.get('country_of_origin', '')
        
        # DEBUG: Print country data
        print(f"[DEBUG] Row {row_num}: country_of_origin='{country_of_origin}' (from tax_calculation_items)", file=sys.stderr)
        
        # Convert country code from 2-letter to 3-digit numeric
        country_code_3digit = country_code_mapping.get(country_of_origin.upper(), '') if country_of_origin else ''
        print(f"[DEBUG] Row {row_num}: country_code_3digit='{country_code_3digit}'", file=sys.stderr)
        
        # Build TANIM (description)
        style = item.get('style', '')
        fabric_content = item.get('fabric_content', '')
        
        # Concatenate: [Style] + [hs_codes.description_tr] + [products.item_description] + [products.fabric_content]
        tanim_parts = []
        if style:
            tanim_parts.append(style)
        if description_tr:
            tanim_parts.append(description_tr)
        if item_description:
            tanim_parts.append(item_description)
        if fabric_content:
            tanim_parts.append(fabric_content)
        
        tanim = ' '.join(tanim_parts)
        
        # Write to columns
        ws[f'A{row_num}'] = tr_hs_code              # GTİP: TR HS CODE
        ws[f'B{row_num}'] = total_value              # KIYMET: Invoice value (Cost × Units)
        
        # MENŞE: Country code as TEXT to preserve leading zeros
        cell_c = ws[f'C{row_num}']
        cell_c.value = country_code_3digit
        cell_c.number_format = '@'  # Format as text to preserve leading zeros
        ws[f'D{row_num}'] = unit                     # MİKTAR CİNS: unit from hs_codes
        ws[f'E{row_num}'] = '1'                      # KAP ADET: Always "1"
        ws[f'F{row_num}'] = 'BI'                     # KAP CİNS: Always "BI"
        ws[f'G{row_num}'] = brand                    # MARKA: brand from products
        ws[f'H{row_num}'] = unit_count               # ADET: units from invoice
        ws[f'I{row_num}'] = 'K1'                     # K1: Always "K1"
        ws[f'J{row_num}'] = '9'                      # SİP TÜRÜ: Always "9"
        ws[f'K{row_num}'] = ''                       # ATR DİĞER: Leave empty
        ws[f'L{row_num}'] = '11'                     # İŞL. NİT: Always "11"
        ws[f'M{row_num}'] = tanim                    # TANIM: Concatenated description
        ws[f'N{row_num}'] = vat_percent              # KDV: vat_percent as percentage
        ws[f'O{row_num}'] = '-'                      # NO: Always "-"
    
    # Save output
    calc = data['calculation']
    output_path = f"/tmp/beyanname_{calc.get('reference', 'export')}_{data.get('timestamp', '')}.xlsx"
    wb.save(output_path)
    
    return output_path

if __name__ == '__main__':
    data = json.loads(sys.stdin.read())
    output_path = export_beyanname_excel(data)
    print(output_path)
